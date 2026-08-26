"""Seed database from ParcelPilot assessment Excel + policy PDFs + demo users."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from apps.accounts.models import Account
from apps.documents.ingestion import seed_documents
from apps.documents.models import DocumentChunk, SourceDocument
from apps.orders.models import Order
from apps.tickets.models import Ticket
from apps.users.models import Role

User = get_user_model()
IST = ZoneInfo("Asia/Kolkata")


def parse_dt(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=IST)
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=IST)
        except ValueError:
            continue
    return None


def as_bool(value) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip() in {"1", "true", "True", "yes"}


def infer_ticket_meta(subject: str, description: str) -> tuple[str, str]:
    text = f"{subject} {description}".lower()
    if "api key" in text or "security" in text:
        return "P1", "security"
    if "all shipment" in text or "http 500" in text:
        return "P1", "outage"
    if "bulk" in text:
        return "P2", "bulk_upload"
    if "swiftship" in text or "booked" in text:
        return "P2", "status_delay"
    if "billing" in text:
        return "P3", "billing"
    if "cancel" in text:
        return "P3", "cancellation"
    return "P3", "general"


class Command(BaseCommand):
    help = "Reset and seed ParcelPilot demo data from Docs for implementation"

    def add_arguments(self, parser):
        parser.add_argument(
            "--skip-documents",
            action="store_true",
            help="Skip PDF ingestion/embeddings",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        docs_dir = Path(settings.DOCS_DIR)
        xlsx = docs_dir / "ParcelPilot_Assessment_Data.xlsx"
        if not xlsx.exists():
            self.stderr.write(f"Missing dataset: {xlsx}")
            return

        self.stdout.write("Clearing existing operational/document data...")
        DocumentChunk.objects.all().delete()
        SourceDocument.objects.all().delete()
        Order.objects.all().delete()
        Ticket.objects.all().delete()
        # Keep accounts if re-run carefully — but for seed reset, recreate
        User.objects.exclude(is_superuser=True).delete()
        Account.objects.all().delete()

        wb = load_workbook(xlsx, data_only=True)

        # Accounts
        ws = wb["accounts"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            Account.objects.create(
                account_code=data["account_id"],
                name=data["account_name"],
                plan=data.get("plan") or "",
                status=data.get("status") or "active",
                csm=data.get("csm") or "",
                contract_file=data.get("contract_file") or "",
                premium_support=as_bool(data.get("premium_support")),
                notes=data.get("notes") or "",
            )
        self.stdout.write(f"Accounts: {Account.objects.count()}")

        # Orders
        ws = wb["orders"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            account = Account.objects.get(account_code=data["account_id"])
            Order.objects.create(
                order_id=data["order_id"],
                account=account,
                carrier=data.get("carrier") or "",
                status=data.get("status") or "",
                booked_at=parse_dt(data.get("booked_at")),
                pickup_window_start=parse_dt(data.get("pickup_window_start")),
                pickup_window_end=parse_dt(data.get("pickup_window_end")),
                pickup_actual_at=parse_dt(data.get("pickup_actual_at")),
                shipment_fee_inr=data.get("shipment_fee_inr") or 0,
                carrier_fault=as_bool(data.get("carrier_fault")),
                customer_fault=as_bool(data.get("customer_fault")),
                cancellation_requested_at=parse_dt(data.get("cancellation_requested_at")),
                notes=data.get("notes") or "",
            )
        self.stdout.write(f"Orders: {Order.objects.count()}")

        # Tickets
        ws = wb["tickets"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            data = dict(zip(headers, row))
            account = Account.objects.get(account_code=data["account_id"])
            subject = data.get("subject") or ""
            description = data.get("description") or ""
            sev, cat = infer_ticket_meta(subject, description)
            Ticket.objects.create(
                ticket_id=data["ticket_id"],
                account=account,
                created_at=parse_dt(data.get("created_at")),
                status=data.get("status") or "open",
                subject=subject,
                description=description,
                channel=data.get("channel") or "",
                assigned_to=data.get("assigned_to") or "",
                last_customer_message_at=parse_dt(data.get("last_customer_message_at")),
                historical_resolution=data.get("historical_resolution") or "",
                severity=sev,
                category=cat,
            )
        self.stdout.write(f"Tickets: {Ticket.objects.count()}")

        # Demo users
        demos = [
            ("northstar@demo.local", "Northstar User", Role.CUSTOMER, "ACCT-001"),
            ("lumenworks@demo.local", "LumenWorks User", Role.CUSTOMER, "ACCT-002"),
            ("support@demo.local", "Internal Support", Role.INTERNAL_SUPPORT, None),
            ("admin@demo.local", "Admin User", Role.ADMIN, None),
        ]
        for email, name, role, acct in demos:
            user = User(
                email=email,
                username=email,
                name=name,
                role=role,
                account=Account.objects.filter(account_code=acct).first() if acct else None,
                is_active=True,
            )
            user.set_password("demo1234")
            user.save()
        self.stdout.write(f"Users: {User.objects.count()}")

        if not options["skip_documents"]:
            self.stdout.write("Ingesting documents (embeddings)...")
            docs = seed_documents(docs_dir)
            self.stdout.write(
                f"Documents: {len(docs)}, chunks: {DocumentChunk.objects.count()}"
            )

        self.stdout.write(self.style.SUCCESS("Seed complete."))
